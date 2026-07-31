"""Wave 1 — Security vehicle entry/exit log: flag gating, permissions, ANPR-or-
manual entry payloads, IN/OUT pairing, currently-inside, idempotent offline
replay, summary counts and XLSX export."""
import uuid

from sqlalchemy import text

from tests.conftest import PHONES, login


async def _enable_flag(db_session):
    await db_session.execute(text("UPDATE settings SET vehicle_log_enabled=true"))
    await db_session.commit()


async def test_feature_flag_off_blocks_real_users(client):
    sec = await login(client, PHONES["w_sec"])
    r = await client.post(
        "/api/vehicles/log",
        json={"plate": "MH12AB1234", "vehicle_type": "truck", "direction": "in"},
        headers=sec,
    )
    assert r.status_code == 403
    assert r.json()["detail"] == {"code": "feature_disabled"}


async def test_non_security_worker_forbidden(client, db_session):
    await _enable_flag(db_session)
    prod_worker = await login(client, PHONES["w_prod1"])
    r = await client.post(
        "/api/vehicles/log",
        json={"plate": "MH12AB1234", "vehicle_type": "truck", "direction": "in"},
        headers=prod_worker,
    )
    assert r.status_code == 403


async def test_in_out_pairing_and_inside(client, db_session):
    await _enable_flag(db_session)
    sec = await login(client, PHONES["w_sec"])

    # IN with messy plate formatting + manual path (no photo)
    r = await client.post(
        "/api/vehicles/log",
        json={
            "plate": "mh-12 ab 4321",
            "vehicle_type": "tractor",
            "direction": "in",
            "driver_name": "Ramesh",
            "purpose": "Cane delivery",
            "gate_zone": "Main Gate",
        },
        headers=sec,
    )
    assert r.status_code == 200, r.text
    in_log = r.json()["log"]
    assert in_log["plate"] == "MH12AB4321"  # normalised
    assert in_log["paired_log_id"] is None

    inside = (await client.get("/api/vehicles/inside", headers=sec)).json()
    assert any(v["id"] == in_log["id"] for v in inside)

    # OUT pairs with the open IN despite different formatting
    r2 = await client.post(
        "/api/vehicles/log",
        json={"plate": "MH12AB4321", "vehicle_type": "tractor", "direction": "out"},
        headers=sec,
    )
    assert r2.status_code == 200
    out_log = r2.json()["log"]
    assert out_log["paired_log_id"] == in_log["id"]

    inside2 = (await client.get("/api/vehicles/inside", headers=sec)).json()
    assert not any(v["id"] == in_log["id"] for v in inside2)


async def test_offline_replay_idempotent(client, db_session):
    await _enable_flag(db_session)
    sec = await login(client, PHONES["w_sec"])
    payload = {
        "plate": "MH14XY9999",
        "vehicle_type": "truck",
        "direction": "in",
        "client_uuid": "outbox-test-uuid-1",
        "anpr_used": True,
    }
    r1 = await client.post("/api/vehicles/log", json=payload, headers=sec)
    r2 = await client.post("/api/vehicles/log", json=payload, headers=sec)
    assert r1.status_code == r2.status_code == 200
    assert r1.json()["log"]["id"] == r2.json()["log"]["id"]
    assert r2.json()["duplicate"] is True
    logs = (await client.get("/api/vehicles/logs", params={"plate": "MH14XY9999"}, headers=sec)).json()
    assert len(logs) == 1
    assert logs[0]["anpr_used"] is True


async def test_validation_rejects_garbage(client, db_session):
    await _enable_flag(db_session)
    sec = await login(client, PHONES["w_sec"])
    r = await client.post(
        "/api/vehicles/log",
        json={"plate": "@@", "vehicle_type": "truck", "direction": "in"},
        headers=sec,
    )
    assert r.status_code == 422
    r = await client.post(
        "/api/vehicles/log",
        json={"plate": "MH12AB1111", "vehicle_type": "spaceship", "direction": "in"},
        headers=sec,
    )
    assert r.status_code == 422


async def test_summary_and_manager_read(client, db_session):
    await _enable_flag(db_session)
    sec = await login(client, PHONES["w_sec"])
    await client.post(
        "/api/vehicles/log",
        json={"plate": "MH09ZZ0001", "vehicle_type": "car", "direction": "in"},
        headers=sec,
    )
    summary = (await client.get("/api/vehicles/summary", headers=sec)).json()
    assert summary["today_in"] >= 1 and summary["currently_inside"] >= 1
    # CGM (rank 2, non-security) can read the register
    cgm = await login(client, PHONES["cgm"])
    logs = (await client.get("/api/vehicles/logs", headers=cgm)).json()
    assert isinstance(logs, list) and len(logs) >= 1


async def test_xlsx_export(client, db_session):
    await _enable_flag(db_session)
    sec = await login(client, PHONES["w_sec"])
    await client.post(
        "/api/vehicles/log",
        json={"plate": "MH20EX0001", "vehicle_type": "bus", "direction": "in"},
        headers=sec,
    )
    cgm = await login(client, PHONES["cgm"])
    r = await client.get("/api/vehicles/export.xlsx", headers=cgm)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert len(r.content) > 500  # a real workbook, not an empty stream


async def test_overstay_sweep_notifies_security_hod(client, db_session):
    """A vehicle IN for >12h with no OUT triggers exactly one HOD alert per day."""
    await _enable_flag(db_session)
    sec = await login(client, PHONES["w_sec"])
    r = await client.post(
        "/api/vehicles/log",
        json={"plate": "MH31OS1111", "vehicle_type": "truck", "direction": "in"},
        headers=sec,
    )
    log_id = r.json()["log"]["id"]
    # age the row 13 hours
    await db_session.execute(
        text("UPDATE vehicle_logs SET logged_at = logged_at - interval '13 hours' WHERE id=:i"),
        {"i": log_id},
    )
    await db_session.commit()

    from app.tasks import _vehicle_overstay_sweep_async

    res1 = await _vehicle_overstay_sweep_async()
    assert res1["overstay_alerts"] >= 1
    res2 = await _vehicle_overstay_sweep_async()
    assert res2["overstay_alerts"] == 0  # dedup: once per vehicle per day

    hod = await login(client, PHONES["sec_mgr"])
    inbox = (await client.get("/api/notifications/mine", headers=hod)).json()
    items = inbox["items"] if isinstance(inbox, dict) else inbox
    assert any(n["type"] == "vehicle_overstay" for n in items)


async def test_webdash_register_flag_off_then_enable_then_empty_then_rows(client, db_session):
    """MD-dashboard regression (2026-07-31 live bug): vehicle_log_enabled was never
    flipped on prod, so every REAL user's register calls returned 403 feature_disabled
    and the webdash stuck on Loading. Covers: flag-off 403s on all three register
    endpoints, the one-click enable path (webdash button = PATCH /admin/settings),
    the clean empty result, and a populated register."""
    await db_session.execute(text("UPDATE settings SET vehicle_log_enabled=false"))
    await db_session.commit()
    cgm = await login(client, PHONES["cgm"])

    for ep in ("/api/vehicles/logs", "/api/vehicles/summary", "/api/vehicles/inside"):
        r = await client.get(ep, headers=cgm)
        assert r.status_code == 403, f"{ep}: {r.status_code}"
        assert r.json()["detail"] == {"code": "feature_disabled"}

    # the webdash "Enable vehicle log" button path — CGM/MD only, audited server-side
    r = await client.patch("/api/admin/settings", json={"vehicle_log_enabled": True}, headers=cgm)
    assert r.status_code == 200 and r.json()["vehicle_log_enabled"] is True

    # EMPTY path: no matching rows must be a clean 200 [] — never an error
    r = await client.get("/api/vehicles/logs?plate=ZZ99XX0000", headers=cgm)
    assert r.status_code == 200 and r.json() == []
    r = await client.get("/api/vehicles/summary", headers=cgm)
    assert r.status_code == 200
    assert set(r.json()) == {"today_in", "today_out", "currently_inside"}

    # POPULATED path: a real gate entry shows up in the register for top mgmt
    sec = await login(client, PHONES["w_sec"])
    r = await client.post(
        "/api/vehicles/log",
        json={"plate": "MH09TT7777", "vehicle_type": "truck", "direction": "in"},
        headers=sec,
    )
    assert r.status_code == 200, r.text
    r = await client.get("/api/vehicles/logs?plate=MH09TT7777", headers=cgm)
    assert r.status_code == 200
    assert len(r.json()) == 1 and r.json()[0]["plate"] == "MH09TT7777"


async def test_export_xlsx_date_range(client, db_session):
    """Time Office monthly gate audit: export accepts a date RANGE and filters by it."""
    await db_session.execute(text("UPDATE settings SET vehicle_log_enabled=true"))
    await db_session.commit()
    cgm = await login(client, PHONES["cgm"])
    sec = await login(client, PHONES["w_sec"])

    r = await client.post(
        "/api/vehicles/log",
        json={"plate": "MH09RNG0001", "vehicle_type": "truck", "direction": "in"},
        headers=sec,
    )
    assert r.status_code == 200, r.text
    new_id = r.json()["log"]["id"]
    # backdate a second entry 40 days into the past (outside this month)
    old_id = str(uuid.uuid4())
    await db_session.execute(
        text(
            "INSERT INTO vehicle_logs (id, plate, vehicle_type, direction, logged_by, logged_at, is_demo) "
            "SELECT :i, 'MH09RNG0002', 'truck', 'in', logged_by, now() - interval '40 days', false "
            "FROM vehicle_logs WHERE id = :n"
        ),
        {"i": old_id, "n": new_id},
    )
    await db_session.commit()

    import io
    from datetime import date as d, timedelta as td
    from openpyxl import load_workbook

    async def export_plates(frm: str, to: str) -> set:
        r = await client.get(f"/api/vehicles/export.xlsx?date_from={frm}&date_to={to}", headers=cgm)
        assert r.status_code == 200, r.text
        ws = load_workbook(io.BytesIO(r.content)).active
        return {row[2] for row in ws.iter_rows(min_row=2, values_only=True)}

    today = d.today()
    month_plates = await export_plates(today.replace(day=1).isoformat(), today.isoformat())
    assert "MH09RNG0001" in month_plates
    assert "MH09RNG0002" not in month_plates  # 40 days old — outside this month

    all_plates = await export_plates((today - td(days=60)).isoformat(), today.isoformat())
    assert {"MH09RNG0001", "MH09RNG0002"} <= all_plates
